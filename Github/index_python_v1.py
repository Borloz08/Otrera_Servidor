# index.py - Python version of the WhatsApp survey bot
import os
import json
import random
from datetime import datetime

from flask import Flask, request
from twilio.rest import Client
import openpyxl

# Twilio credentials must be set as environment variables
ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID')
AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN')
WHATSAPP_NUMBER = os.getenv('TWILIO_WHATSAPP_NUMBER')  # e.g. 'whatsapp:+14155238886'

client = Client(ACCOUNT_SID, AUTH_TOKEN)
app = Flask(__name__)

# Load states and municipalities from Excel
wb = openpyxl.load_workbook('Estados-Municipios.xlsx')
ws = wb.active
estado_municipios = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    estado = str(row[0]).strip() if row[0] else ''
    municipio = str(row[1]).strip() if row[1] else ''
    if not estado or not municipio:
        continue
    estado_municipios.setdefault(estado, []).append(municipio)

lista_estados = list(estado_municipios.keys())

preguntas_generales = [
    "¿Cuántos años tiene? Debe tener entre 18 y 100 años para responder esta encuesta.",
    "Por favor indique su género\n1) Hombre\n2) Mujer\n3) Otro",
    "Escoja el estado de México en el que vive:",
    "Seleccione el municipio donde vive:",
    "¿Cuál es su estado civil?\n1) Soltero\n2) Casado\n3) Divorciado\n4) Viudo\n5) Cónyuge\n99) NC",
    "¿Hasta qué año aprobó en la escuela?\n1) Ninguno\n2) Preescolar\n3) Primaria\n4) Secundaria\n5) Carrera técnica con secundaria terminada\n6) Normal básica\n7) Preparatoria o bachillerato\n8) Carrera técnica con preparatoria\n9) Licenciatura o profesional\n10) Maestría o doctorado\n99) NC"
]

preguntas_grupo = {
    'tratamiento': {
        7: "El gobierno de Estados Unidos ha amenazado con imponer nuevos aranceles a México, los cuales afectarían mucho la economía de nuestro país, a menos que México acepte negociar un nuevo acuerdo de seguridad. En una escala de 0 al 10 (siendo 0 totalmente en contra y 10 totalmente a favor) ¿qué tan en contra o a favor está de que México negocie un acuerdo de seguridad con Estados Unidos?",
        '8_variantes': [
            "¿Y qué tan en contra o a favor estaría de que una de las condiciones de dicho acuerdo fuera que México se comprometa a extraditar los narcotraficantes más buscados, es decir enviarlos a Estados Unidos?",
            "¿Y qué tan en contra o a favor estaría de que una de las condiciones de dicho acuerdo fuera que México se comprometa a permitir que Estados Unidos use drones para vigilar a los carteles mexicanos?",
            "¿Y qué tan en contra o a favor estaría de que una de las condiciones de dicho acuerdo fuera que México se comprometa a permitir la entrada de tropas estadounidenses en territorio mexicano para asegurar la frontera?"
        ]
    },
    'control': {
        7: "En una escala de 0 al 10 (siendo 0 totalmente en contra y 10 totalmente a favor) ¿Qué tan en contra o a favor está de que México negocie un acuerdo de seguridad con Estados Unidos?",
        '8_variantes': [
            "¿Y qué tan en contra o a favor estaría de que se negocie un acuerdo en el que México se comprometa a extraditar los narcotraficantes más buscados, es decir enviarlos a Estados Unidos?",
            "¿Y qué tan en contra o a favor estaría de que se negocie un acuerdo en el que México se comprometa a permitir que Estados Unidos use drones para atacar a los carteles mexicanos?",
            "¿Y qué tan en contra o a favor estaría de que se negocie un acuerdo en el que México se comprometa a permitir la entrada de tropas estadounidenses en territorio mexicano para asegurar la frontera?"
        ]
    }
}

# Contacts to receive the survey
contactos = {
    '50670600922': {'nombre': 'Andres', 'grupo': 'tratamiento'}
}

estados = {}
respuestas_globales = {}


def generar_timestamp():
    now = datetime.now()
    return now.strftime('%Y-%m-%d_%H-%M-%S')


def guardar_json(path):
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(respuestas_globales, fh, indent=2, ensure_ascii=False)


def guardar_excel(path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = [
        'Nombre', 'Telefono', 'Grupo',
        *[f'P{i+1}' for i in range(len(preguntas_generales))],
        'P7', 'P8.A', 'P8.B', 'P8.C'
    ]
    ws.append(headers)

    for phone, registro in respuestas_globales.items():
        fila = [registro['nombre'], phone, registro['grupo']]
        for i in range(len(preguntas_generales)):
            fila.append(registro['respuestas'][i] if i < len(registro['respuestas']) else '')
        fila.append(registro['respuestas'][len(preguntas_generales)] if len(registro['respuestas']) > len(preguntas_generales) else '')
        fila.extend(registro['p8'])
        ws.append(fila)

    wb.save(path)


def send_message(to, text):
    client.messages.create(
        from_=f'whatsapp:{WHATSAPP_NUMBER}',
        to=f'whatsapp:{to}',
        body=text
    )


def send_initial_questions():
    for numero, info in contactos.items():
        nombre = info['nombre']
        grupo = info['grupo']
        variantes = preguntas_grupo[grupo]['8_variantes'][:]
        random.shuffle(variantes)
        orden8 = list(enumerate(variantes))
        estados[numero] = {
            'indice': 0,
            'respuestas': [],
            'nombre': nombre,
            'grupo': grupo,
            'ordenP8': orden8,
            'estadoSeleccionado': '',
            'municipios': []
        }
        send_message(numero, f"P1. {nombre}, {preguntas_generales[0]}")


def validar_respuesta(indice, texto, estado=None):
    try:
        num = int(texto)
    except ValueError:
        num = None
    if indice == 0:
        return num is not None and 18 <= num <= 100
    if indice == 1:
        return num in {1, 2, 3}
    if indice == 4:
        return num in {1, 2, 3, 4, 5, 99}
    if indice == 5:
        return num in {1,2,3,4,5,6,7,8,9,10,99}
    if indice in {6,7,8,9}:
        return num is not None and 0 <= num <= 10
    if indice == 2:
        return num is not None and 1 <= num <= len(lista_estados)
    if indice == 3 and estado is not None:
        return num is not None and 1 <= num <= len(estado['municipios'])
    return True


@app.route('/webhook', methods=['POST'])
def whatsapp_webhook():
    from_number = request.values.get('From', '').replace('whatsapp:', '')
    body = request.values.get('Body', '').strip()
    estado = estados.get(from_number)
    if not estado:
        return ('', 204)

    indice = estado['indice']
    if not validar_respuesta(indice, body, estado):
        send_message(from_number, '❌ Respuesta inválida. Por favor responda con una opción válida.')
        return ('', 204)

    if indice == 2:
        index = int(body)
        estado['estadoSeleccionado'] = lista_estados[index-1]
        estado['municipios'] = estado_municipios[estado['estadoSeleccionado']]
        estado['respuestas'].append(estado['estadoSeleccionado'])
    elif indice == 3:
        index = int(body)
        municipio = estado['municipios'][index-1]
        estado['respuestas'].append(municipio)
    else:
        estado['respuestas'].append(body)

    estado['indice'] += 1

    if estado['indice'] == 2:
        texto = f"P3. {estado['nombre']}, escoja el estado de México en el que vive:\n"
        for idx, est in enumerate(lista_estados, 1):
            texto += f"{idx}) {est}\n"
        send_message(from_number, texto)
    elif estado['indice'] == 3:
        texto = f"P4. {estado['nombre']}, seleccione el municipio donde vive:\n"
        for idx, m in enumerate(estado['municipios'], 1):
            texto += f"{idx}) {m}\n"
        send_message(from_number, texto)
    elif estado['indice'] < len(preguntas_generales):
        num = estado['indice'] + 1
        texto = f"P{num}. {estado['nombre']}, {preguntas_generales[estado['indice']]}"
        send_message(from_number, texto)
    elif estado['indice'] == len(preguntas_generales):
        texto = f"P7. {estado['nombre']}, {preguntas_grupo[estado['grupo']][7]}"
        send_message(from_number, texto)
    elif len(preguntas_generales) < estado['indice'] <= len(preguntas_generales)+2:
        idx8 = estado['indice'] - (len(preguntas_generales) + 1)
        pnum = estado['indice'] + 1
        texto = f"P{pnum}. {estado['nombre']}, {estado['ordenP8'][idx8][1]}"
        send_message(from_number, texto)
    else:
        send_message(from_number, f"✅ ¡Gracias {estado['nombre']} por completar la encuesta!")

        respuestas_ordenadas = [''] * 3
        for index, (orig_idx, _) in enumerate(estado['ordenP8']):
            respuestas_ordenadas[orig_idx] = estado['respuestas'][len(preguntas_generales)+1+index]

        respuestas_globales[from_number] = {
            'nombre': estado['nombre'],
            'grupo': estado['grupo'],
            'respuestas': estado['respuestas'][:len(preguntas_generales)+1],
            'p8': respuestas_ordenadas
        }

        ts = generar_timestamp()
        guardar_json(f'encuestas_{ts}.json')
        guardar_excel(f'encuestas_{ts}.xlsx')

        estados.pop(from_number, None)

    return ('', 204)


if __name__ == '__main__':
    send_initial_questions()
    app.run(host='0.0.0.0', port=5000)
