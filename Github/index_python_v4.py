# index.py - Python version of the WhatsApp survey bot
import os
import json
import random
from datetime import datetime

from flask import Flask, request
from twilio.rest import Client
import openpyxl

# Twilio credentials must be set as environment variables
ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID')
AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN')
WHATSAPP_NUMBER = os.getenv('TWILIO_WHATSAPP_NUMBER')  # e.g. 'whatsapp:+14155238886'

client = Client(ACCOUNT_SID, AUTH_TOKEN)
app = Flask(__name__)

# Limit of unique phone numbers allowed per QR code
N_PHONES_PER_CODE = 5
CODIGOS_FILE = 'codigos.json'


def cargar_codigos():
    try:
        with open(CODIGOS_FILE, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except FileNotFoundError:
        return {}


def guardar_codigos():
    with open(CODIGOS_FILE, 'w', encoding='utf-8') as fh:
        json.dump(codigos_usados, fh, indent=2, ensure_ascii=False)


codigos_usados = cargar_codigos()

# Load states and municipalities from Excel
wb = openpyxl.load_workbook('Estados-Municipios.xlsx')
ws = wb.active
estado_municipios = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    estado = str(row[0]).strip() if row[0] else ''
    municipio = str(row[1]).strip() if row[1] else ''
    if not estado or not municipio:
        continue
    estado_municipios.setdefault(estado, []).append(municipio)

lista_estados = list(estado_municipios.keys())

# Approximate coordinates (latitude, longitude) for each Mexican state
estado_centros = {
    "Aguascalientes": (21.8853, -102.2916),
    "Baja California": (32.6245, -115.4523),
    "Baja California Sur": (24.1426, -110.3128),
    "Campeche": (19.845, -90.523),
    "Chiapas": (16.7524, -93.116),
    "Chihuahua": (28.632, -106.069),
    "Ciudad de México": (19.4326, -99.1332),
    "Coahuila": (25.4232, -101.0053),
    "Colima": (19.2433, -103.7241),
    "Durango": (24.0277, -104.6532),
@@ -117,58 +137,58 @@ def haversine(lat1, lon1, lat2, lon2):
    return R * c


def detectar_estado_por_ubicacion(lat, lon):
    mejor_estado = None
    mejor_dist = float('inf')
    for estado, (elat, elon) in estado_centros.items():
        dist = haversine(lat, lon, elat, elon)
        if dist < mejor_dist:
            mejor_dist = dist
            mejor_estado = estado
    return mejor_estado


def guardar_json(path):
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(respuestas_globales, fh, indent=2, ensure_ascii=False)


def guardar_excel(path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = [
        'Nombre', 'Telefono', 'Grupo',
        'Nombre', 'Telefono', 'Grupo', 'Codigo',
        *[f'P{i+1}' for i in range(len(preguntas_generales))],
        'P7', 'P8.A', 'P8.B', 'P8.C'
    ]
    ws.append(headers)

    for phone, registro in respuestas_globales.items():
        fila = [registro['nombre'], phone, registro['grupo']]
        fila = [registro['nombre'], phone, registro['grupo'], registro['codigo']]
        for i in range(len(preguntas_generales)):
            fila.append(registro['respuestas'][i] if i < len(registro['respuestas']) else '')
        fila.append(registro['respuestas'][len(preguntas_generales)] if len(registro['respuestas']) > len(preguntas_generales) else '')
        fila.extend(registro['p8'])
        ws.append(fila)

    wb.save(path)


def send_message(to, text):
    client.messages.create(
        from_=f'whatsapp:{WHATSAPP_NUMBER}',
        to=f'whatsapp:{to}',
        body=text
    )


def send_initial_questions():
    for numero, info in contactos.items():
        nombre = info['nombre']
        grupo = info['grupo']
        variantes = preguntas_grupo[grupo]['8_variantes'][:]
        random.shuffle(variantes)
        orden8 = list(enumerate(variantes))
        estados[numero] = {
@@ -191,50 +211,79 @@ def validar_respuesta(indice, texto, estado=None, lat=None, lon=None):
    if indice == 0:
        return num is not None and 18 <= num <= 100
    if indice == 1:
        return num in {1, 2, 3}
    if indice == 4:
        return num in {1, 2, 3, 4, 5, 99}
    if indice == 5:
        return num in {1,2,3,4,5,6,7,8,9,10,99}
    if indice in {6,7,8,9}:
        return num is not None and 0 <= num <= 10
    if indice == 2:
        return lat is not None and lon is not None
    if indice == 3 and estado is not None:
        return num is not None and 1 <= num <= len(estado['municipios'])
    return True


@app.route('/webhook', methods=['POST'])
def whatsapp_webhook():
    from_number = request.values.get('From', '').replace('whatsapp:', '')
    body = request.values.get('Body', '').strip()
    lat = request.values.get('Latitude')
    lon = request.values.get('Longitude')
    estado = estados.get(from_number)
    if not estado:
        codigo = body
        usados = codigos_usados.get(codigo, [])
        if from_number not in usados and len(usados) >= N_PHONES_PER_CODE:
            send_message(from_number, '❌ Este código ya alcanzó el máximo de teléfonos permitidos.')
            return ('', 204)

        perfil = request.values.get('ProfileName', from_number)
        grupo = random.choice(['tratamiento', 'control'])
        variantes = preguntas_grupo[grupo]['8_variantes'][:]
        random.shuffle(variantes)
        orden8 = list(enumerate(variantes))

        estados[from_number] = {
            'indice': 0,
            'respuestas': [],
            'nombre': perfil,
            'grupo': grupo,
            'ordenP8': orden8,
            'estadoSeleccionado': '',
            'municipios': [],
            'codigo': codigo
        }

        if from_number not in usados:
            usados.append(from_number)
            codigos_usados[codigo] = usados
            guardar_codigos()

        send_message(from_number, f"P1. {perfil}, {preguntas_generales[0]}")
        return ('', 204)

    indice = estado['indice']
    if not validar_respuesta(indice, body, estado, lat, lon):
        send_message(from_number, '❌ Respuesta inválida. Por favor responda con una opción válida.')
        return ('', 204)

    if indice == 2:
        try:
            lat_f = float(lat)
            lon_f = float(lon)
        except (TypeError, ValueError):
            send_message(from_number, '❌ No se recibió una ubicación válida. Inténtelo de nuevo.')
            return ('', 204)

        estado_nombre = detectar_estado_por_ubicacion(lat_f, lon_f)
        if not estado_nombre:
            send_message(from_number, '❌ No se pudo determinar su estado. Intente enviando nuevamente su ubicación.')
            return ('', 204)

        estado['estadoSeleccionado'] = estado_nombre
        estado['municipios'] = estado_municipios.get(estado_nombre, [])
        estado['respuestas'].append(estado_nombre)
    elif indice == 3:
        index = int(body)
@@ -253,41 +302,42 @@ def whatsapp_webhook():
        for idx, m in enumerate(estado['municipios'], 1):
            texto += f"{idx}) {m}\n"
        send_message(from_number, texto)
    elif estado['indice'] < len(preguntas_generales):
        num = estado['indice'] + 1
        texto = f"P{num}. {estado['nombre']}, {preguntas_generales[estado['indice']]}"
        send_message(from_number, texto)
    elif estado['indice'] == len(preguntas_generales):
        texto = f"P7. {estado['nombre']}, {preguntas_grupo[estado['grupo']][7]}"
        send_message(from_number, texto)
    elif len(preguntas_generales) < estado['indice'] <= len(preguntas_generales)+2:
        idx8 = estado['indice'] - (len(preguntas_generales) + 1)
        pnum = estado['indice'] + 1
        texto = f"P{pnum}. {estado['nombre']}, {estado['ordenP8'][idx8][1]}"
        send_message(from_number, texto)
    else:
        send_message(from_number, f"✅ ¡Gracias {estado['nombre']} por completar la encuesta!")

        respuestas_ordenadas = [''] * 3
        for index, (orig_idx, _) in enumerate(estado['ordenP8']):
            respuestas_ordenadas[orig_idx] = estado['respuestas'][len(preguntas_generales)+1+index]

        respuestas_globales[from_number] = {
            'nombre': estado['nombre'],
            'grupo': estado['grupo'],
            'codigo': estado['codigo'],
            'respuestas': estado['respuestas'][:len(preguntas_generales)+1],
            'p8': respuestas_ordenadas
        }

        ts = generar_timestamp()
        guardar_json(f'encuestas_{ts}.json')
        guardar_excel(f'encuestas_{ts}.xlsx')

        estados.pop(from_number, None)

    return ('', 204)


if __name__ == '__main__':
    send_initial_questions()
    app.run(host='0.0.0.0', port=5000)
